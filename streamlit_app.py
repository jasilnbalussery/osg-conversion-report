import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="OSG Target Achievement Report", layout="wide")

st.title("📊 OSG Target Achievement Report Generator")

# Target percentages
TARGET_MAP = {
    'TV': 5,
    'MICROWAVE OVEN': 5,
    'REFRIGERATOR': 2,
    'AC': 1,
    'WASHING MACHINE': 3,
    'SMALL APPLIANCE': 2
}


def normalize_category(category):
    if pd.isna(category):
        return None
    cat_upper = str(category).upper().strip()
    for target_cat in TARGET_MAP.keys():
        if cat_upper == target_cat:
            return target_cat
    if 'WASHING MACHINE' in cat_upper or 'WASHING' in cat_upper or 'WASHER' in cat_upper:
        return 'WASHING MACHINE'
    elif 'WM' == cat_upper or cat_upper.startswith('WM ') or ' WM' in cat_upper or cat_upper.endswith(' WM'):
        return 'WASHING MACHINE'
    elif 'TELEVISION' in cat_upper or cat_upper == 'TV' or cat_upper.startswith('TV ') or ' TV' in cat_upper or cat_upper.endswith(' TV'):
        return 'TV'
    elif 'MICROWAVE' in cat_upper or ('MICRO' in cat_upper and 'WAVE' in cat_upper):
        return 'MICROWAVE OVEN'
    elif 'OVEN' in cat_upper and 'MICROWAVE' not in cat_upper:
        return 'MICROWAVE OVEN'
    elif 'REFRIGERATOR' in cat_upper or 'FRIDGE' in cat_upper or 'REFG' in cat_upper:
        return 'REFRIGERATOR'
    elif 'REF' == cat_upper or cat_upper.startswith('REF ') or ' REF' in cat_upper or cat_upper.endswith(' REF'):
        return 'REFRIGERATOR'
    elif 'AIR CONDITIONER' in cat_upper or 'AIRCONDITIONER' in cat_upper or 'AIR CONDITION' in cat_upper:
        return 'AC'
    elif 'AC' == cat_upper or cat_upper.startswith('AC ') or ' AC' in cat_upper or cat_upper.endswith(' AC'):
        return 'AC'
    elif 'SMALL APPLIANCE' in cat_upper or 'SMALL APP' in cat_upper:
        return 'SMALL APPLIANCE'
    elif 'SA' == cat_upper or cat_upper.startswith('SA ') or ' SA' in cat_upper or cat_upper.endswith(' SA'):
        return 'SMALL APPLIANCE'
    return None


def is_future_store(branch_name):
    """Check if a branch/store is a Future Store"""
    if pd.isna(branch_name):
        return False
    name_upper = str(branch_name).upper().strip()
    # Match various Future store naming patterns
    if 'FUTURE' in name_upper:
        return True
    if 'FRL' in name_upper:
        return True
    if 'FRETAIL' in name_upper or 'F RETAIL' in name_upper:
        return True
    return False


def process_data(product_df, osg_df):
    product_cols = {'rbm': None, 'branch': None, 'category': None, 'sold_price': None}
    osg_cols = {'branch': None, 'category': None, 'sold_price': None}

    for col in product_df.columns:
        col_upper = col.upper()
        if 'RBM' in col_upper:
            product_cols['rbm'] = col
        elif 'BRANCH' in col_upper:
            product_cols['branch'] = col
        elif 'CATEGORY' in col_upper:
            product_cols['category'] = col
        elif 'SOLD PRICE' in col_upper:
            product_cols['sold_price'] = col
        elif 'TAXABLE VALUE' in col_upper and product_cols['sold_price'] is None:
            product_cols['sold_price'] = col

    for col in osg_df.columns:
        col_upper = col.upper()
        if 'STORE NAME' in col_upper or 'BRANCH' in col_upper:
            osg_cols['branch'] = col
        elif 'CATEGORY' in col_upper:
            osg_cols['category'] = col
        elif 'SOLD PRICE' in col_upper:
            osg_cols['sold_price'] = col

    missing_product = [k for k, v in product_cols.items() if v is None]
    missing_osg = [k for k, v in osg_cols.items() if v is None]

    if missing_product or missing_osg:
        st.error(f"Missing columns - Product: {missing_product}, OSG: {missing_osg}")
        return None

    product_clean = product_df[[
        product_cols['rbm'], product_cols['branch'],
        product_cols['category'], product_cols['sold_price']
    ]].copy()
    product_clean.columns = ['RBM', 'Branch', 'Category', 'Product_Sold_Price']

    osg_clean = osg_df[[
        osg_cols['branch'], osg_cols['category'], osg_cols['sold_price']
    ]].copy()
    osg_clean.columns = ['Branch', 'Category', 'OSG_Sold_Price']

    product_clean['Category_Normalized'] = product_clean['Category'].apply(normalize_category)
    osg_clean['Category_Normalized'] = osg_clean['Category'].apply(normalize_category)

    product_clean = product_clean[product_clean['Category_Normalized'].notna()].copy()
    osg_clean = osg_clean[osg_clean['Category_Normalized'].notna()].copy()

    product_agg = product_clean.groupby(['RBM', 'Branch', 'Category_Normalized']).agg({
        'Product_Sold_Price': 'sum'
    }).reset_index()
    product_agg.columns = ['RBM', 'Branch', 'Category', 'Product_Sold_Price']

    osg_agg = osg_clean.groupby(['Branch', 'Category_Normalized']).agg({
        'OSG_Sold_Price': 'sum'
    }).reset_index()
    osg_agg.columns = ['Branch', 'Category', 'OSG_Sold_Price']

    merged = product_agg.merge(osg_agg, on=['Branch', 'Category'], how='left')
    merged['OSG_Sold_Price'] = merged['OSG_Sold_Price'].fillna(0)
    merged['Target_%'] = merged['Category'].map(TARGET_MAP)

    merged['Value_Conversion_%'] = np.where(
        merged['Product_Sold_Price'] > 0,
        (merged['OSG_Sold_Price'] / merged['Product_Sold_Price'] * 100).round(2), 0
    )

    merged['Need_to_Achieve_Target_%'] = np.maximum(
        (merged['Target_%'] - merged['Value_Conversion_%']).round(2), 0
    )

    merged['Target_Achieved'] = merged['Value_Conversion_%'] >= merged['Target_%']

    # ── Flag Future Stores ──
    merged['Is_Future_Store'] = merged['Branch'].apply(is_future_store)

    final_df = merged[[
        'RBM', 'Branch', 'Category', 'Product_Sold_Price', 'OSG_Sold_Price',
        'Value_Conversion_%', 'Target_%', 'Need_to_Achieve_Target_%',
        'Target_Achieved', 'Is_Future_Store'
    ]].copy()

    return final_df


def build_store_overview(df, future_only=False):
    """Build store-wise overall conversion summary"""
    work_df = df.copy()
    if future_only:
        work_df = work_df[work_df['Is_Future_Store']].copy()

    if len(work_df) == 0:
        return pd.DataFrame()

    store_summary = work_df.groupby(['Branch', 'RBM']).agg({
        'Product_Sold_Price': 'sum',
        'OSG_Sold_Price': 'sum'
    }).reset_index()

    store_summary['Value_Conversion_%'] = np.where(
        store_summary['Product_Sold_Price'] > 0,
        (store_summary['OSG_Sold_Price'] / store_summary['Product_Sold_Price'] * 100).round(2), 0
    )

    wtd = work_df.groupby('Branch').apply(
        lambda g: np.average(g['Target_%'], weights=g['Product_Sold_Price'])
        if g['Product_Sold_Price'].sum() > 0 else 0
    ).reset_index(name='Wtd_Target_%')

    store_summary = store_summary.merge(wtd, on='Branch')
    store_summary['Need_to_Achieve_%'] = np.maximum(
        (store_summary['Wtd_Target_%'] - store_summary['Value_Conversion_%']).round(2), 0
    )

    cats_met = work_df[work_df['Target_Achieved']].groupby('Branch').size().reset_index(name='Categories_Target_Met')
    cats_total = work_df.groupby('Branch').size().reset_index(name='Total_Categories')

    store_summary = store_summary.merge(cats_met, on='Branch', how='left')
    store_summary = store_summary.merge(cats_total, on='Branch', how='left')
    store_summary['Categories_Target_Met'] = store_summary['Categories_Target_Met'].fillna(0).astype(int)

    store_summary = store_summary.sort_values('Value_Conversion_%', ascending=False).reset_index(drop=True)
    store_summary['Rank'] = range(1, len(store_summary) + 1)

    return store_summary


def build_top_bottom_analysis(df, future_store_overview):
    """Build category-wise analysis for top 5 and bottom 5 FUTURE stores"""

    if len(future_store_overview) == 0:
        return [], [], pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), \
            ["No Future Stores found in data. Check branch naming."]

    n_top = min(5, len(future_store_overview))
    n_bot = min(5, len(future_store_overview))

    top5_stores = future_store_overview.head(n_top)['Branch'].tolist()
    bottom5_stores = future_store_overview.tail(n_bot)['Branch'].tolist()

    # Filter original data for these stores
    top5_data = df[df['Branch'].isin(top5_stores)].copy()
    bottom5_data = df[df['Branch'].isin(bottom5_stores)].copy()

    # Category-wise aggregation for top 5
    top5_cat = top5_data.groupby('Category').agg({
        'Product_Sold_Price': 'sum',
        'OSG_Sold_Price': 'sum',
        'Target_%': 'first'
    }).reset_index()
    top5_cat['Value_Conversion_%'] = np.where(
        top5_cat['Product_Sold_Price'] > 0,
        (top5_cat['OSG_Sold_Price'] / top5_cat['Product_Sold_Price'] * 100).round(2), 0
    )
    top5_cat['Need_to_Achieve_%'] = np.maximum(
        (top5_cat['Target_%'] - top5_cat['Value_Conversion_%']).round(2), 0
    )

    # Category-wise aggregation for bottom 5
    bot5_cat = bottom5_data.groupby('Category').agg({
        'Product_Sold_Price': 'sum',
        'OSG_Sold_Price': 'sum',
        'Target_%': 'first'
    }).reset_index()
    bot5_cat['Value_Conversion_%'] = np.where(
        bot5_cat['Product_Sold_Price'] > 0,
        (bot5_cat['OSG_Sold_Price'] / bot5_cat['Product_Sold_Price'] * 100).round(2), 0
    )
    bot5_cat['Need_to_Achieve_%'] = np.maximum(
        (bot5_cat['Target_%'] - bot5_cat['Value_Conversion_%']).round(2), 0
    )

    # Comparison table
    comparison = top5_cat[['Category', 'Value_Conversion_%']].rename(
        columns={'Value_Conversion_%': 'Top5_Conversion_%'}
    ).merge(
        bot5_cat[['Category', 'Value_Conversion_%']].rename(
            columns={'Value_Conversion_%': 'Bottom5_Conversion_%'}
        ), on='Category', how='outer'
    )
    comparison = comparison.merge(
        pd.DataFrame({'Category': list(TARGET_MAP.keys()), 'Target_%': list(TARGET_MAP.values())}),
        on='Category', how='left'
    )
    comparison['Top5_Conversion_%'] = comparison['Top5_Conversion_%'].fillna(0)
    comparison['Bottom5_Conversion_%'] = comparison['Bottom5_Conversion_%'].fillna(0)
    comparison['Gap_%'] = (comparison['Top5_Conversion_%'] - comparison['Bottom5_Conversion_%']).round(2)
    comparison['Top5_vs_Target'] = (comparison['Top5_Conversion_%'] - comparison['Target_%']).round(2)
    comparison['Bottom5_vs_Target'] = (comparison['Bottom5_Conversion_%'] - comparison['Target_%']).round(2)

    insights = generate_insights(df, future_store_overview,
                                 top5_data, bottom5_data, comparison)

    return top5_stores, bottom5_stores, top5_cat, bot5_cat, comparison, insights


def generate_insights(df, future_overview, top5_data, bottom5_data, comparison):
    """Generate text-based insights for FUTURE stores"""
    insights = []

    # Count future stores
    future_df = df[df['Is_Future_Store']].copy()
    total_future = future_df['Branch'].nunique()
    total_all = df['Branch'].nunique()

    insights.append(
        f"1. SCOPE: Analysing {total_future} Future Stores out of "
        f"{total_all} total stores ({total_future / total_all * 100:.1f}%)"
    )

    # 2. Future stores overall conversion
    total_prod = future_df['Product_Sold_Price'].sum()
    total_osg = future_df['OSG_Sold_Price'].sum()
    overall_conv = (total_osg / total_prod * 100) if total_prod > 0 else 0
    insights.append(
        f"2. FUTURE STORES OVERALL: Product Sales = ₹{total_prod:,.0f}, "
        f"OSG Sales = ₹{total_osg:,.0f}, Conversion = {overall_conv:.2f}%"
    )

    # 3. Top 5 vs Bottom 5 avg
    if len(future_overview) >= 5:
        top5_avg = future_overview.head(5)['Value_Conversion_%'].mean()
        bot5_avg = future_overview.tail(5)['Value_Conversion_%'].mean()
        insights.append(
            f"3. TOP 5 Future Stores Avg = {top5_avg:.2f}% vs "
            f"BOTTOM 5 Future Stores Avg = {bot5_avg:.2f}% "
            f"(Gap = {top5_avg - bot5_avg:.2f}%)"
        )

    # 4. Best category in top 5
    if len(comparison) > 0:
        best_top = comparison.loc[comparison['Top5_Conversion_%'].idxmax()]
        insights.append(
            f"4. BEST CATEGORY in Top 5 Future Stores: {best_top['Category']} "
            f"at {best_top['Top5_Conversion_%']:.2f}% (Target: {best_top['Target_%']:.0f}%)"
        )

    # 5. Worst category in bottom 5
    if len(comparison) > 0:
        worst_bot = comparison.loc[comparison['Bottom5_Conversion_%'].idxmin()]
        insights.append(
            f"5. WEAKEST CATEGORY in Bottom 5 Future Stores: {worst_bot['Category']} "
            f"at {worst_bot['Bottom5_Conversion_%']:.2f}% (Target: {worst_bot['Target_%']:.0f}%)"
        )

    # 6. Biggest gap
    if len(comparison) > 0:
        max_gap = comparison.loc[comparison['Gap_%'].idxmax()]
        insights.append(
            f"6. BIGGEST GAP between Top & Bottom Future Stores: {max_gap['Category']} "
            f"— Top5: {max_gap['Top5_Conversion_%']:.2f}%, "
            f"Bottom5: {max_gap['Bottom5_Conversion_%']:.2f}%, Gap: {max_gap['Gap_%']:.2f}%"
        )

    # 7. Categories where even top 5 miss target
    if len(comparison) > 0:
        below = comparison[comparison['Top5_vs_Target'] < 0]
        if len(below) > 0:
            cats = ", ".join(below['Category'].tolist())
            insights.append(f"7. ALERT — Even Top 5 Future Stores BELOW target in: {cats}")
        else:
            insights.append("7. Top 5 Future Stores meet targets in ALL categories ✅")

    # 8. Categories where bottom 5 meet target
    if len(comparison) > 0:
        above = comparison[comparison['Bottom5_vs_Target'] >= 0]
        if len(above) > 0:
            cats = ", ".join(above['Category'].tolist())
            insights.append(f"8. POSITIVE — Bottom 5 Future Stores MEET target in: {cats}")
        else:
            insights.append("8. Bottom 5 Future Stores are below target in ALL categories ⚠️")

    # 9. Best single future store
    if len(future_overview) > 0:
        best = future_overview.iloc[0]
        insights.append(
            f"9. BEST FUTURE STORE: {best['Branch']} (RBM: {best['RBM']}) "
            f"— Conversion: {best['Value_Conversion_%']:.2f}%"
        )

    # 10. Worst single future store
    if len(future_overview) > 0:
        worst = future_overview.iloc[-1]
        insights.append(
            f"10. NEEDS ATTENTION: {worst['Branch']} (RBM: {worst['RBM']}) "
            f"— Conversion: {worst['Value_Conversion_%']:.2f}%"
        )

    # 11. How many future stores meet target
    if len(future_overview) > 0:
        stores_above = len(future_overview[
            future_overview['Value_Conversion_%'] >= future_overview['Wtd_Target_%']
        ])
        total_s = len(future_overview)
        insights.append(
            f"11. FUTURE STORES MEETING TARGET: {stores_above}/{total_s} "
            f"({stores_above / total_s * 100:.1f}%)"
        )

    # 12. Recommendation
    if len(comparison) > 0:
        worst_gap_cats = comparison.nlargest(2, 'Gap_%')['Category'].tolist()
        insights.append(
            f"12. RECOMMENDATION: Focus OSG push in Bottom 5 Future Stores on "
            f"{', '.join(worst_gap_cats)} — these have the largest performance gap vs top stores"
        )

    return insights


def create_excel_report(df, future_keyword):
    """Create full Excel report with all sheets"""

    wb = Workbook()
    wb.remove(wb.active)

    # ── Styles ──
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    green_font = Font(color='006100')
    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    red_font = Font(color='9C0006')

    top5_row_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
    top5_row_font = Font(color='375623', bold=True)
    bot5_row_fill = PatternFill(start_color='FCE4EC', end_color='FCE4EC', fill_type='solid')
    bot5_row_font = Font(color='B71C1C', bold=True)

    section_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    section_font = Font(bold=True, color='FFFFFF', size=12)

    insight_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    insight_font_style = Font(color='7F6000', size=11)

    future_label_fill = PatternFill(start_color='DAEEF3', end_color='DAEEF3', fill_type='solid')
    future_label_font = Font(bold=True, color='1F4E78', size=11)

    def apply_header(ws, row, headers, start_col=1):
        for col_idx, h in enumerate(headers, start=start_col):
            cell = ws.cell(row=row, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

    def write_row(ws, row, values, start_col=1, fmt_map=None):
        for col_idx, v in enumerate(values, start=start_col):
            cell = ws.cell(row=row, column=col_idx, value=v)
            cell.border = thin_border
            cell.alignment = Alignment(
                horizontal='left' if col_idx <= 2 else 'right', vertical='center')
            if fmt_map and col_idx in fmt_map:
                cell.number_format = fmt_map[col_idx]

    def add_section_header(ws, row, text, end_col=8):
        cell = ws.cell(row=row, column=1, value=text)
        cell.font = section_font
        cell.fill = section_fill
        for c in range(2, end_col + 1):
            ws.cell(row=row, column=c).fill = section_fill
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=end_col)

    # ── Build data ──
    all_store_overview = build_store_overview(df, future_only=False)
    future_store_overview = build_store_overview(df, future_only=True)
    top5_stores, bottom5_stores, top5_cat, bot5_cat, comparison, insights = \
        build_top_bottom_analysis(df, future_store_overview)

    # ============================================================
    # SHEET 1: RBM Summary
    # ============================================================
    ws_summary = wb.create_sheet('RBM Summary')
    ws_summary['A1'] = 'OSG TARGET ACHIEVEMENT — RBM SUMMARY'
    ws_summary['A1'].font = Font(bold=True, size=16, color='1F4E78')
    ws_summary.merge_cells('A1:F1')
    ws_summary['A1'].alignment = Alignment(horizontal='center')

    rbm_summary = df.groupby('RBM').agg({
        'Product_Sold_Price': 'sum', 'OSG_Sold_Price': 'sum'
    }).reset_index()
    rbm_summary['Value_Conversion_%'] = np.where(
        rbm_summary['Product_Sold_Price'] > 0,
        (rbm_summary['OSG_Sold_Price'] / rbm_summary['Product_Sold_Price'] * 100).round(2), 0)
    rbm_target = df.groupby('RBM').apply(
        lambda g: np.average(g['Target_%'], weights=g['Product_Sold_Price'])
        if g['Product_Sold_Price'].sum() > 0 else 0
    ).reset_index(name='Wtd_Target_%')
    rbm_summary = rbm_summary.merge(rbm_target, on='RBM')
    rbm_summary['Need_to_Achieve_%'] = np.maximum(
        (rbm_summary['Wtd_Target_%'] - rbm_summary['Value_Conversion_%']).round(2), 0)

    s_headers = ['RBM', 'Product Sold Price', 'OSG Sold Price',
                 'Value Conversion (%)', 'Wtd. Target (%)', 'Need to Achieve (%)']
    apply_header(ws_summary, 3, s_headers)

    fmt = {2: '₹#,##0.00', 3: '#,##0', 4: '0.00"%"', 5: '0.00"%"', 6: '0.00"%"'}
    for r, row_data in enumerate(rbm_summary.values, start=4):
        write_row(ws_summary, r, row_data, fmt_map=fmt)
        need_cell = ws_summary.cell(row=r, column=6)
        if row_data[3] >= row_data[4]:
            need_cell.fill = green_fill; need_cell.font = green_font
        else:
            need_cell.fill = red_fill; need_cell.font = red_font

    for c in range(1, 7):
        ws_summary.column_dimensions[get_column_letter(c)].width = 25

    # ============================================================
    # SHEET 2: Store-wise Overview (ALL stores)
    # ============================================================
    ws_store = wb.create_sheet('Store Overview - All')
    ws_store['A1'] = 'STORE-WISE OVERALL CONVERSION — ALL STORES'
    ws_store['A1'].font = Font(bold=True, size=16, color='1F4E78')
    ws_store.merge_cells('A1:J1')
    ws_store['A1'].alignment = Alignment(horizontal='center')

    ws_store['A2'] = '🔷 Future Store rows highlighted in blue'
    ws_store['A2'].font = Font(italic=True, size=10, color='666666')
    ws_store.merge_cells('A2:J2')

    st_headers = ['Rank', 'Branch', 'RBM', 'Store Type', 'Product Sold Price',
                  'OSG Sold Price', 'Value Conversion (%)', 'Wtd. Target (%)',
                  'Need to Achieve (%)', 'Categories Met']
    apply_header(ws_store, 4, st_headers)

    st_fmt = {5: '₹#,##0.00', 6: '#,##0', 7: '0.00"%"', 8: '0.00"%"', 9: '0.00"%"'}

    for r_idx, (_, row) in enumerate(all_store_overview.iterrows(), start=5):
        is_future = is_future_store(row['Branch'])
        store_type = 'FUTURE' if is_future else 'OTHER'
        vals = [
            int(row['Rank']), row['Branch'], row['RBM'], store_type,
            row['Product_Sold_Price'], row['OSG_Sold_Price'],
            row['Value_Conversion_%'], row['Wtd_Target_%'],
            row['Need_to_Achieve_%'],
            f"{int(row['Categories_Target_Met'])}/{int(row['Total_Categories'])}"
        ]
        write_row(ws_store, r_idx, vals, fmt_map=st_fmt)

        # Highlight Future stores with blue tint
        if is_future:
            for c in range(1, 11):
                cell = ws_store.cell(row=r_idx, column=c)
                cell.fill = future_label_fill
                cell.font = future_label_font

    widths = [8, 35, 20, 14, 22, 20, 22, 18, 22, 16]
    for i, w in enumerate(widths, start=1):
        ws_store.column_dimensions[get_column_letter(i)].width = w

    # ============================================================
    # SHEET 3: Future Store Overview (FUTURE only)
    # ============================================================
    ws_future = wb.create_sheet('Future Store Ranking')
    ws_future['A1'] = 'FUTURE STORES — CONVERSION RANKING'
    ws_future['A1'].font = Font(bold=True, size=16, color='1F4E78')
    ws_future.merge_cells('A1:I1')
    ws_future['A1'].alignment = Alignment(horizontal='center')

    future_count = len(future_store_overview)
    ws_future['A2'] = (
        f'Total Future Stores: {future_count} | '
        f'🟢 Top 5 highlighted green | 🔴 Bottom 5 highlighted red'
    )
    ws_future['A2'].font = Font(italic=True, size=10, color='666666')
    ws_future.merge_cells('A2:I2')

    fs_headers = ['Rank', 'Branch', 'RBM', 'Product Sold Price', 'OSG Sold Price',
                  'Value Conversion (%)', 'Wtd. Target (%)', 'Need to Achieve (%)',
                  'Categories Met']
    apply_header(ws_future, 4, fs_headers)

    fs_fmt = {4: '₹#,##0.00', 5: '#,##0', 6: '0.00"%"', 7: '0.00"%"', 8: '0.00"%"'}

    for r_idx, (_, row) in enumerate(future_store_overview.iterrows(), start=5):
        rank = row['Rank']
        vals = [
            int(rank), row['Branch'], row['RBM'],
            row['Product_Sold_Price'], row['OSG_Sold_Price'],
            row['Value_Conversion_%'], row['Wtd_Target_%'],
            row['Need_to_Achieve_%'],
            f"{int(row['Categories_Target_Met'])}/{int(row['Total_Categories'])}"
        ]
        write_row(ws_future, r_idx, vals, fmt_map=fs_fmt)

        is_top5 = rank <= 5
        is_bot5 = rank > future_count - 5

        if is_top5:
            for c in range(1, 10):
                ws_future.cell(row=r_idx, column=c).fill = top5_row_fill
                ws_future.cell(row=r_idx, column=c).font = top5_row_font
        elif is_bot5:
            for c in range(1, 10):
                ws_future.cell(row=r_idx, column=c).fill = bot5_row_fill
                ws_future.cell(row=r_idx, column=c).font = bot5_row_font

    fs_widths = [8, 35, 20, 22, 20, 22, 18, 22, 16]
    for i, w in enumerate(fs_widths, start=1):
        ws_future.column_dimensions[get_column_letter(i)].width = w

    # ============================================================
    # SHEET 4: Top 5 vs Bottom 5 FUTURE Stores Analysis
    # ============================================================
    ws_tb = wb.create_sheet('Top5 vs Bottom5 Future')
    ws_tb['A1'] = 'TOP 5 vs BOTTOM 5 FUTURE STORES — CATEGORY ANALYSIS & INSIGHTS'
    ws_tb['A1'].font = Font(bold=True, size=14, color='1F4E78')
    ws_tb.merge_cells('A1:H1')
    ws_tb['A1'].alignment = Alignment(horizontal='center')

    ws_tb['A2'] = '⚡ Analysis restricted to FUTURE STORES only'
    ws_tb['A2'].font = Font(bold=True, italic=True, size=11, color='C00000')
    ws_tb.merge_cells('A2:H2')

    current_row = 4

    # ── Section A: Top 5 Future Stores ──
    add_section_header(ws_tb, current_row,
                       '🏆 TOP 5 FUTURE STORES (Highest Conversion)')
    current_row += 1

    top5_detail_headers = ['#', 'Branch', 'RBM', 'Product Sold Price',
                           'OSG Sold Price', 'Value Conversion (%)']
    apply_header(ws_tb, current_row, top5_detail_headers)
    current_row += 1

    top5_overview = future_store_overview[
        future_store_overview['Branch'].isin(top5_stores)
    ]
    for i, (_, row) in enumerate(top5_overview.iterrows(), start=1):
        vals = [i, row['Branch'], row['RBM'],
                row['Product_Sold_Price'], row['OSG_Sold_Price'],
                row['Value_Conversion_%']]
        write_row(ws_tb, current_row, vals,
                  fmt_map={4: '₹#,##0.00', 5: '#,##0', 6: '0.00"%"'})
        for c in range(1, 7):
            ws_tb.cell(row=current_row, column=c).fill = top5_row_fill
            ws_tb.cell(row=current_row, column=c).font = top5_row_font
        current_row += 1

    current_row += 1

    # ── Section B: Bottom 5 Future Stores ──
    add_section_header(ws_tb, current_row,
                       '⚠️ BOTTOM 5 FUTURE STORES (Lowest Conversion)')
    current_row += 1

    apply_header(ws_tb, current_row, top5_detail_headers)
    current_row += 1

    bot5_overview = future_store_overview[
        future_store_overview['Branch'].isin(bottom5_stores)
    ]
    for i, (_, row) in enumerate(bot5_overview.iterrows(), start=1):
        vals = [i, row['Branch'], row['RBM'],
                row['Product_Sold_Price'], row['OSG_Sold_Price'],
                row['Value_Conversion_%']]
        write_row(ws_tb, current_row, vals,
                  fmt_map={4: '₹#,##0.00', 5: '#,##0', 6: '0.00"%"'})
        for c in range(1, 7):
            ws_tb.cell(row=current_row, column=c).fill = bot5_row_fill
            ws_tb.cell(row=current_row, column=c).font = bot5_row_font
        current_row += 1

    current_row += 2

    # ── Section C: Category Comparison ──
    add_section_header(ws_tb, current_row,
                       '📊 CATEGORY COMPARISON: TOP 5 vs BOTTOM 5 FUTURE STORES')
    current_row += 1

    comp_headers = ['Category', 'Target %', 'Top 5 Conversion (%)',
                    'Bottom 5 Conversion (%)', 'Gap (Top5 − Bottom5) %',
                    'Top5 vs Target', 'Bottom5 vs Target', 'Verdict']
    apply_header(ws_tb, current_row, comp_headers)
    current_row += 1

    for _, crow in comparison.iterrows():
        if crow['Top5_vs_Target'] >= 0 and crow['Bottom5_vs_Target'] >= 0:
            verdict = '✅ Both meet target'
        elif crow['Top5_vs_Target'] >= 0 and crow['Bottom5_vs_Target'] < 0:
            verdict = '⚠️ Bottom 5 need improvement'
        elif crow['Top5_vs_Target'] < 0 and crow['Bottom5_vs_Target'] < 0:
            verdict = '🚨 Both below target'
        else:
            verdict = '🔍 Review needed'

        vals = [crow['Category'], crow['Target_%'],
                crow['Top5_Conversion_%'], crow['Bottom5_Conversion_%'],
                crow['Gap_%'], crow['Top5_vs_Target'],
                crow['Bottom5_vs_Target'], verdict]
        write_row(ws_tb, current_row, vals,
                  fmt_map={2: '0.00"%"', 3: '0.00"%"', 4: '0.00"%"',
                           5: '0.00"%"', 6: '0.00"%"', 7: '0.00"%"'})

        gap_cell = ws_tb.cell(row=current_row, column=5)
        if crow['Gap_%'] > 3:
            gap_cell.fill = red_fill; gap_cell.font = red_font
        elif crow['Gap_%'] > 0:
            gap_cell.fill = PatternFill(
                start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')

        t5_cell = ws_tb.cell(row=current_row, column=6)
        if crow['Top5_vs_Target'] >= 0:
            t5_cell.fill = green_fill; t5_cell.font = green_font
        else:
            t5_cell.fill = red_fill; t5_cell.font = red_font

        b5_cell = ws_tb.cell(row=current_row, column=7)
        if crow['Bottom5_vs_Target'] >= 0:
            b5_cell.fill = green_fill; b5_cell.font = green_font
        else:
            b5_cell.fill = red_fill; b5_cell.font = red_font

        current_row += 1

    current_row += 2

    # ── Section D: Category Detail — Top 5 ──
    add_section_header(ws_tb, current_row,
                       '📋 CATEGORY DETAIL — TOP 5 FUTURE STORES (Aggregated)')
    current_row += 1

    cat_headers = ['Category', 'Product Sold Price', 'OSG Sold Price',
                   'Value Conversion (%)', 'Target %', 'Need to Achieve (%)']
    apply_header(ws_tb, current_row, cat_headers)
    current_row += 1

    for _, crow in top5_cat.iterrows():
        vals = [crow['Category'], crow['Product_Sold_Price'], crow['OSG_Sold_Price'],
                crow['Value_Conversion_%'], crow['Target_%'], crow['Need_to_Achieve_%']]
        write_row(ws_tb, current_row, vals,
                  fmt_map={2: '₹#,##0.00', 3: '#,##0',
                           4: '0.00"%"', 5: '0.00"%"', 6: '0.00"%"'})
        need_cell = ws_tb.cell(row=current_row, column=6)
        if crow['Need_to_Achieve_%'] == 0:
            need_cell.fill = green_fill; need_cell.font = green_font
        else:
            need_cell.fill = red_fill; need_cell.font = red_font
        current_row += 1

    current_row += 1

    # ── Section E: Category Detail — Bottom 5 ──
    add_section_header(ws_tb, current_row,
                       '📋 CATEGORY DETAIL — BOTTOM 5 FUTURE STORES (Aggregated)')
    current_row += 1

    apply_header(ws_tb, current_row, cat_headers)
    current_row += 1

    for _, crow in bot5_cat.iterrows():
        vals = [crow['Category'], crow['Product_Sold_Price'], crow['OSG_Sold_Price'],
                crow['Value_Conversion_%'], crow['Target_%'], crow['Need_to_Achieve_%']]
        write_row(ws_tb, current_row, vals,
                  fmt_map={2: '₹#,##0.00', 3: '#,##0',
                           4: '0.00"%"', 5: '0.00"%"', 6: '0.00"%"'})
        need_cell = ws_tb.cell(row=current_row, column=6)
        if crow['Need_to_Achieve_%'] == 0:
            need_cell.fill = green_fill; need_cell.font = green_font
        else:
            need_cell.fill = red_fill; need_cell.font = red_font
        current_row += 1

    current_row += 2

    # ── Section F: Insights ──
    add_section_header(ws_tb, current_row,
                       '💡 KEY INSIGHTS & RECOMMENDATIONS — FUTURE STORES')
    current_row += 1

    for insight in insights:
        cell = ws_tb.cell(row=current_row, column=1, value=insight)
        cell.font = insight_font_style
        cell.fill = insight_fill
        ws_tb.merge_cells(start_row=current_row, start_column=1,
                          end_row=current_row, end_column=8)
        cell.alignment = Alignment(wrap_text=True, vertical='center')
        ws_tb.row_dimensions[current_row].height = 30
        current_row += 1

    tb_widths = [22, 25, 25, 22, 22, 22, 22, 35]
    for i, w in enumerate(tb_widths, start=1):
        ws_tb.column_dimensions[get_column_letter(i)].width = w

    # ============================================================
    # SHEETS 5+: Individual RBM sheets
    # ============================================================
    rbms = sorted(df['RBM'].unique())

    for rbm in rbms:
        rbm_data = df[df['RBM'] == rbm].copy()
        rbm_data = rbm_data.drop(['RBM', 'Target_Achieved', 'Is_Future_Store'], axis=1)
        rbm_data = rbm_data.sort_values(['Branch', 'Category'])

        sheet_name = str(rbm)[:31]
        ws = wb.create_sheet(sheet_name)

        ws['A1'] = f'Target Achievement Report - {rbm}'
        ws['A1'].font = Font(bold=True, size=14, color='1F4E78')
        ws.merge_cells('A1:G1')
        ws['A1'].alignment = Alignment(horizontal='center')

        headers = ['Branch', 'Category', 'Product Sold Price', 'OSG Sold Price',
                   'Value Conversion (%)', 'Target %', 'Need to Achieve Target (%)']
        apply_header(ws, 3, headers)

        r_fmt = {3: '₹#,##0.00', 4: '#,##0', 5: '0.00"%"', 6: '0.00"%"', 7: '0.00"%"'}
        for row_idx, row_data in enumerate(rbm_data.values, start=4):
            write_row(ws, row_idx, row_data, fmt_map=r_fmt)
            need_cell = ws.cell(row=row_idx, column=7)
            conv_val = row_data[4]
            target_val = row_data[5]
            if conv_val >= target_val:
                need_cell.fill = green_fill; need_cell.font = green_font
            else:
                need_cell.fill = red_fill; need_cell.font = red_font

        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 22
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 22
        ws.column_dimensions['F'].width = 14
        ws.column_dimensions['G'].width = 28

        last_row = len(rbm_data) + 3
        sr = last_row + 2
        ws.cell(row=sr, column=1, value='TOTAL').font = Font(bold=True)
        ws.cell(row=sr, column=3, value=f'=SUM(C4:C{last_row})')
        ws.cell(row=sr, column=3).number_format = '₹#,##0.00'
        ws.cell(row=sr, column=3).font = Font(bold=True)
        ws.cell(row=sr, column=4, value=f'=SUM(D4:D{last_row})')
        ws.cell(row=sr, column=4).number_format = '#,##0'
        ws.cell(row=sr, column=4).font = Font(bold=True)
        ws.cell(row=sr, column=5, value=f'=IF(C{sr}>0,(D{sr}/C{sr})*100,0)')
        ws.cell(row=sr, column=5).number_format = '0.00"%"'
        ws.cell(row=sr, column=5).font = Font(bold=True)

    return wb


def main():
    st.sidebar.header("📁 Upload Files")

    product_file = st.sidebar.file_uploader(
        "Upload Product File (Excel)", type=['xlsx', 'xls'], key='product')
    osg_file = st.sidebar.file_uploader(
        "Upload OSG File (Excel)", type=['xlsx', 'xls'], key='osg')

    # ── Future store keyword ──
    st.sidebar.markdown("---")
    st.sidebar.header("🏪 Future Store Filter")
    future_keyword = st.sidebar.text_input(
        "Keyword to identify Future Stores (in branch name)",
        value="FUTURE",
        help="Branches containing this keyword will be treated as Future Stores"
    )

    if product_file and osg_file:
        with st.spinner("Loading files..."):
            product_df = pd.read_excel(product_file)
            osg_df = pd.read_excel(osg_file)

        st.success(f"✅ Files loaded! Product: {len(product_df)} rows, OSG: {len(osg_df)} rows")

        with st.expander("📋 Detected Columns"):
            c1, c2 = st.columns(2)
            with c1:
                st.write("**Product File Columns:**", product_df.columns.tolist())
            with c2:
                st.write("**OSG File Columns:**", osg_df.columns.tolist())

        with st.expander("🔍 Category Analysis"):
            c1, c2 = st.columns(2)
            product_cat_col = next(
                (c for c in product_df.columns if 'CATEGORY' in c.upper()), None)
            osg_cat_col = next(
                (c for c in osg_df.columns if 'CATEGORY' in c.upper()), None)
            with c1:
                if product_cat_col:
                    st.write("**Product Categories:**")
                    st.dataframe(product_df[product_cat_col].value_counts().head(20),
                                 use_container_width=True)
            with c2:
                if osg_cat_col:
                    st.write("**OSG Categories:**")
                    st.dataframe(osg_df[osg_cat_col].value_counts().head(20),
                                 use_container_width=True)

        if st.button("🚀 Generate Report", type="primary"):
            with st.spinner("Processing data..."):
                processed_df = process_data(product_df, osg_df)

            if processed_df is not None:
                st.success(f"✅ {len(processed_df)} records processed")

                # ── Stats ──
                future_count = processed_df[processed_df['Is_Future_Store']]['Branch'].nunique()
                total_stores = processed_df['Branch'].nunique()

                st.info(
                    f"🏪 **Future Stores detected: {future_count}** out of "
                    f"{total_stores} total stores "
                    f"(keyword: '{future_keyword}')"
                )

                if future_count == 0:
                    st.warning(
                        "⚠️ No Future Stores found! Check if your branch names "
                        f"contain '{future_keyword}'. You can change the keyword in the sidebar."
                    )

                    # Show sample branch names for debugging
                    sample_branches = processed_df['Branch'].unique()[:20]
                    st.write("**Sample branch names in your data:**")
                    st.write(list(sample_branches))

                # ── Build analysis ──
                all_store_overview = build_store_overview(processed_df, future_only=False)
                future_store_overview = build_store_overview(processed_df, future_only=True)
                top5_stores, bottom5_stores, top5_cat, bot5_cat, comparison, insights = \
                    build_top_bottom_analysis(processed_df, future_store_overview)

                # ── Tabs ──
                tab1, tab2, tab3, tab4, tab5 = st.tabs([
                    "📊 RBM Summary", "🏪 All Stores",
                    "🔷 Future Stores", "🏆 Top5 vs Bottom5", "💡 Insights"
                ])

                with tab1:
                    st.subheader("RBM-wise Data Preview")
                    st.dataframe(processed_df.head(30), use_container_width=True)
                    c1, c2, c3, c4 = st.columns(4)
                    with c1:
                        st.metric("Total Product Sales",
                                  f"₹{processed_df['Product_Sold_Price'].sum():,.0f}")
                    with c2:
                        st.metric("Total OSG Sales",
                                  f"₹{processed_df['OSG_Sold_Price'].sum():,.0f}")
                    with c3:
                        oc = (processed_df['OSG_Sold_Price'].sum() /
                              processed_df['Product_Sold_Price'].sum() * 100)
                        st.metric("Overall Conversion", f"{oc:.2f}%")
                    with c4:
                        st.metric("Targets Met",
                                  f"{processed_df['Target_Achieved'].sum()}/{len(processed_df)}")

                with tab2:
                    st.subheader("🏪 All Stores Ranking")
                    if len(all_store_overview) > 0:
                        display_cols = ['Rank', 'Branch', 'RBM', 'Product_Sold_Price',
                                        'OSG_Sold_Price', 'Value_Conversion_%',
                                        'Categories_Target_Met', 'Total_Categories']
                        st.dataframe(all_store_overview[display_cols],
                                     use_container_width=True)

                with tab3:
                    st.subheader("🔷 Future Stores Ranking")
                    if len(future_store_overview) > 0:
                        display_cols = ['Rank', 'Branch', 'RBM', 'Product_Sold_Price',
                                        'OSG_Sold_Price', 'Value_Conversion_%',
                                        'Categories_Target_Met', 'Total_Categories']

                        def highlight_top_bottom(row):
                            if row['Rank'] <= 5:
                                return ['background-color: #C6EFCE; color: #006100; '
                                        'font-weight: bold'] * len(row)
                            elif row['Rank'] > len(future_store_overview) - 5:
                                return ['background-color: #FFC7CE; color: #9C0006; '
                                        'font-weight: bold'] * len(row)
                            return [''] * len(row)

                        st.dataframe(
                            future_store_overview[display_cols].style.apply(
                                highlight_top_bottom, axis=1),
                            use_container_width=True
                        )

                        c1, c2 = st.columns(2)
                        with c1:
                            f_prod = future_store_overview['Product_Sold_Price'].sum()
                            f_osg = future_store_overview['OSG_Sold_Price'].sum()
                            f_conv = (f_osg / f_prod * 100) if f_prod > 0 else 0
                            st.metric("Future Stores Conversion", f"{f_conv:.2f}%")
                        with c2:
                            st.metric("Future Stores Count", future_count)
                    else:
                        st.warning("No Future Stores found.")

                with tab4:
                    if len(future_store_overview) > 0:
                        st.subheader("🏆 Top 5 Future Stores")
                        st.dataframe(
                            future_store_overview[
                                future_store_overview['Branch'].isin(top5_stores)
                            ][['Branch', 'RBM', 'Value_Conversion_%']],
                            use_container_width=True
                        )

                        st.subheader("⚠️ Bottom 5 Future Stores")
                        st.dataframe(
                            future_store_overview[
                                future_store_overview['Branch'].isin(bottom5_stores)
                            ][['Branch', 'RBM', 'Value_Conversion_%']],
                            use_container_width=True
                        )

                        st.subheader("📊 Category Comparison (Future Stores Only)")
                        st.dataframe(comparison, use_container_width=True)
                    else:
                        st.warning("No Future Stores found for analysis.")

                with tab5:
                    st.subheader("💡 Key Insights — Future Stores")
                    for insight in insights:
                        st.info(insight)

                # ── Generate Excel ──
                with st.spinner("Creating Excel report..."):
                    wb = create_excel_report(processed_df, future_keyword)
                    buffer = BytesIO()
                    wb.save(buffer)
                    buffer.seek(0)

                st.success("✅ Report generated!")

                st.download_button(
                    label="📥 Download Complete Excel Report",
                    data=buffer,
                    file_name="OSG_Target_Achievement_Report.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

                st.subheader("📑 Excel Sheets in Report:")
                rbms = sorted(processed_df['RBM'].unique())
                st.markdown(f"""
                | # | Sheet | Description |
                |---|-------|-------------|
                | 1 | **RBM Summary** | RBM-wise conversion vs weighted target |
                | 2 | **Store Overview - All** | All stores ranked with Future stores highlighted 🔷 |
                | 3 | **Future Store Ranking** | Only Future stores ranked — Top 5 🟢 Bottom 5 🔴 |
                | 4 | **Top5 vs Bottom5 Future** | Category comparison + insights for Future stores only |
                | 5–{4 + len(rbms)} | **{len(rbms)} RBM Sheets** | {', '.join(rbms)} |
                """)

    else:
        st.info("👆 Upload both Product and OSG files")
        st.markdown(f"""
        ### 📊 Report Structure (5 types of sheets):

        | Sheet | What it contains |
        |-------|------------------|
        | **RBM Summary** | RBM-wise conversion vs weighted target |
        | **Store Overview - All** | Every store ranked, Future stores highlighted 🔷 |
        | **Future Store Ranking** | Only Future stores, Top 5 🟢 / Bottom 5 🔴 |
        | **Top5 vs Bottom5 Future** | Category analysis + comparison + 12 auto-insights |
        | **RBM Sheets** | Branch × Category detail per RBM |

        ### 🏪 Future Store Detection:
        - Stores with **"FUTURE"** in branch name are auto-detected
        - You can change the keyword in the sidebar
        - Also matches **FRL**, **FRETAIL** patterns

        ### 🎯 Targets:
        TV: 5% · Microwave: 5% · Refrigerator: 2% · AC: 1% · WM: 3% · Small Appliance: 2%

        ### 🏆 Top5 vs Bottom5 Analysis includes:
        1. Top 5 & Bottom 5 Future store lists with RBM & conversion
        2. Category-by-category comparison table with verdicts
        3. Aggregated category detail for each group
        4. **12 auto-generated insights** covering gaps, alerts & recommendations
        """)


if __name__ == "__main__":
    main()
